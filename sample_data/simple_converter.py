"""
Simple Universal Excel Converter - Fixed Version
Handles any Excel file structure and converts to standard format
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def convert_any_excel_to_standard(input_path: str, output_path: str) -> bool:
    """
    Convert any Excel file to standard student format
    
    Args:
        input_path: Path to input Excel file
        output_path: Path to output converted file
    
    Returns:
        bool: True if conversion successful
    """
    try:
        # Read the Excel file
        df = pd.read_excel(input_path, dtype=str)
        logger.info(f"Loaded Excel file: {len(df)} rows, {len(df.columns)} columns")
        
        # Detect structure type
        structure_type = detect_structure(df)
        logger.info(f"Detected structure: {structure_type}")
        
        # Convert based on structure type
        if structure_type == "question_bank":
            result_df = convert_question_bank(df)
        elif structure_type == "student_format":
            result_df = convert_student_format(df)
        elif structure_type == "tabular":
            result_df = convert_tabular(df)
        else:
            result_df = convert_generic(df)
        
        # Save with formatting
        save_with_formatting(result_df, output_path)
        logger.info(f"Successfully converted to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Conversion failed: {e}")
        return False

def detect_structure(df: pd.DataFrame) -> str:
    """Detect the structure type of the Excel file"""
    columns = [str(col).strip().lower() for col in df.columns]
    
    # Check for question bank format
    question_bank_indicators = ['questiontext', 'question', 'optiona', 'optionb', 'optionc', 'optiond']
    if any(indicator in columns for indicator in question_bank_indicators):
        return "question_bank"
    
    # Check for student format
    student_indicators = ['name', 'enrollmentno', 'roll', 'rollno']
    if any(indicator in columns for indicator in student_indicators):
        return "student_format"
    
    # Check for tabular format
    tabular_indicators = ['s.no', 'serial', 'question', 'answer']
    if any(indicator in columns for indicator in tabular_indicators):
        return "tabular"
    
    return "generic"

def convert_question_bank(df: pd.DataFrame) -> pd.DataFrame:
    """Convert question bank format to student format"""
    # Find columns
    question_col = find_column(df, ['questiontext', 'question', 'qtext', 'text'])
    option_a_col = find_column(df, ['optiona', 'a', 'choicea', 'opta'])
    option_b_col = find_column(df, ['optionb', 'b', 'choiceb', 'optb'])
    option_c_col = find_column(df, ['optionc', 'c', 'choicec', 'optc'])
    option_d_col = find_column(df, ['optiond', 'd', 'choiced', 'optd'])
    answer_col = find_column(df, ['correctanswer', 'answer', 'correct', 'key'])
    
    # Create student data
    students = []
    base_names = ["Aarav Sharma", "Priya Patel", "Rohit Verma", "Sneha Gupta", "Arjun Singh"]
    
    # Limit questions
    max_questions = min(20, len(df))
    questions_df = df.head(max_questions)
    
    for i, name in enumerate(base_names[:5]):
        student_data = {
            'Name': name,
            'EnrollmentNo': f'EN2024{str(i+1).zfill(3)}'
        }
        
        # Add questions
        questions = questions_df.copy()
        if i > 0:  # Shuffle for different students
            questions = questions.sample(frac=1).reset_index(drop=True)
        
        for j, (_, row) in enumerate(questions.iterrows(), 1):
            student_data[f'Q{j}'] = str(row[question_col]).strip()
            student_data[f'Q{j}_A'] = str(row[option_a_col]).strip()
            student_data[f'Q{j}_B'] = str(row[option_b_col]).strip()
            student_data[f'Q{j}_C'] = str(row[option_c_col]).strip() if option_c_col else ''
            student_data[f'Q{j}_D'] = str(row[option_d_col]).strip() if option_d_col else ''
            student_data[f'Q{j}_Answer'] = str(row[answer_col]).strip().upper() if answer_col else ''
        
        students.append(student_data)
    
    return pd.DataFrame(students)

def convert_student_format(df: pd.DataFrame) -> pd.DataFrame:
    """Convert simple student format to standard format"""
    result_df = df.copy()
    
    # Map common column variations
    column_map = {
        'name': ['name', 'student', 'studentname', 's_name'],
        'enrollmentno': ['enrollmentno', 'enrollment', 'roll', 'rollno', 'roll_no', 'id']
    }
    
    mapped_df = pd.DataFrame()
    for target_col, possible_cols in column_map.items():
        for possible_col in possible_cols:
            for actual_col in df.columns:
                if str(actual_col).lower() == possible_col.lower():
                    mapped_df[target_col] = df[actual_col]
                    break
    
    # Check if we have required columns
    if 'name' not in mapped_df.columns or 'enrollmentno' not in mapped_df.columns:
        logger.warning("Could not map required columns, using generic conversion")
        return convert_generic(df)
    
    # Copy other columns as-is
    for col in df.columns:
        if col not in mapped_df.columns:
            mapped_df[col] = df[col]
    
    return mapped_df

def convert_tabular(df: pd.DataFrame) -> pd.DataFrame:
    """Convert tabular format to student format"""
    # Try to detect if first column contains student names
    first_col = df.columns[0]
    if looks_like_student_names(df[first_col]):
        return convert_from_tabular_student_data(df)
    else:
        # Try to convert as question bank
        return convert_question_bank(df)

def convert_generic(df: pd.DataFrame) -> pd.DataFrame:
    """Generic conversion as last resort"""
    logger.info("Using generic conversion")
    
    result_df = pd.DataFrame()
    
    # Try to find name column
    name_col = find_column(df, ['name', 'student', 'person', 'user'])
    if name_col:
        result_df['Name'] = df[name_col]
    else:
        # Create generic names
        result_df['Name'] = [f"Student {i+1}" for i in range(len(df))]
    
    # Try to find ID column
    id_col = find_column(df, ['id', 'roll', 'enrollment', 'number'])
    if id_col:
        result_df['EnrollmentNo'] = df[id_col]
    else:
        # Generate IDs
        result_df['EnrollmentNo'] = [f"STU{str(i+1).zfill(4)}" for i in range(len(df))]
    
    # Add placeholder questions
    for i in range(1, 11):  # 10 questions
        result_df[f'Q{i}'] = f"Question {i} (Please edit with actual content)"
        result_df[f'Q{i}_A'] = "Option A"
        result_df[f'Q{i}_B'] = "Option B"
        result_df[f'Q{i}_C'] = "Option C"
        result_df[f'Q{i}_D'] = "Option D"
        result_df[f'Q{i}_Answer'] = "A"
    
    return result_df

def convert_from_tabular_student_data(df: pd.DataFrame) -> pd.DataFrame:
    """Convert tabular data where first column is student names"""
    student_names = df.iloc[:, 0].dropna()
    
    question_data = []
    for i, name in enumerate(student_names):
        student_data = {'Name': name, 'EnrollmentNo': f'STU{str(i+1).zfill(4)}'}
        
        # Process remaining columns as questions
        col_idx = 1
        question_num = 1
        
        while col_idx < len(df.columns) and question_num <= 20:
            if col_idx + 3 < len(df.columns):
                student_data[f'Q{question_num}'] = str(df.iloc[i, col_idx]).strip()
                student_data[f'Q{question_num}_A'] = str(df.iloc[i, col_idx + 1]).strip()
                student_data[f'Q{question_num}_B'] = str(df.iloc[i, col_idx + 2]).strip()
                student_data[f'Q{question_num}_C'] = str(df.iloc[i, col_idx + 3]).strip()
                student_data[f'Q{question_num}_D'] = ""
                student_data[f'Q{question_num}_Answer'] = "A"
                col_idx += 4
            else:
                student_data[f'Q{question_num}'] = f"Question {question_num}"
                student_data[f'Q{question_num}_A'] = "Option A"
                student_data[f'Q{question_num}_B'] = "Option B"
                student_data[f'Q{question_num}_C'] = "Option C"
                student_data[f'Q{question_num}_D'] = "Option D"
                student_data[f'Q{question_num}_Answer'] = "A"
                col_idx += 1
            
            question_num += 1
        
        question_data.append(student_data)
    
    return pd.DataFrame(question_data)

def find_column(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """Find the best matching column from possible names"""
    for possible_name in possible_names:
        for col in df.columns:
            if str(col).strip().lower() == possible_name.lower():
                return col
    return None

def looks_like_student_names(series: pd.Series) -> bool:
    """Check if a series looks like it contains student names"""
    sample = series.head(10).dropna()
    if len(sample) == 0:
        return False
    
    name_like_count = 0
    for value in sample:
        val_str = str(value).strip()
        if (len(val_str) > 3 and len(val_str) < 50 and 
            any(c.isalpha() for c in val_str) and
            not any(c.isdigit() for c in val_str.split()[0] if val_str.split())):
            name_like_count += 1
    
    return name_like_count / len(sample) > 0.6

def save_with_formatting(df: pd.DataFrame, output_path: str) -> None:
    """Save DataFrame with proper Excel formatting"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Write headers
        headers = list(df.columns)
        header_fill = PatternFill("solid", fgColor="1F3A8F")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        ws.row_dimensions[1].height = 22
        
        # Write data
        row_fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]
        
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            ws.row_dimensions[row_idx].height = 20
            fill = row_fills[(row_idx - 2) % 2]
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            if col_idx == 1:  # Name column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            elif col_idx == 2:  # Enrollment column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        wb.save(output_path)
        logger.info(f"File saved with formatting: {output_path}")
        
    except Exception as e:
        logger.error(f"Failed to save with formatting: {e}")
        # Fallback to simple save
        df.to_excel(output_path, index=False)

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python simple_converter.py <input_file> [output_file]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else f"converted_{input_file}"
    
    success = convert_any_excel_to_standard(input_file, output_file)
    
    if success:
        print(f"✅ Successfully converted {input_file} to {output_file}")
    else:
        print(f"❌ Failed to convert {input_file}")
        sys.exit(1)
