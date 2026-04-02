"""
Excel File Converter for Student PDF Generator
Converts various Excel formats to the required format for the Student PDF Generator.

Usage: python convert_excel.py input_file.xlsx output_file.xlsx
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os
from pathlib import Path

def detect_columns(df):
    """Detect and map column names to the expected format"""
    detected = {
        'name': None,
        'enrollment': None,
        'questions': []
    }
    
    # Common variations for name column
    name_variations = ['name', 'student_name', 'studentname', 'student', 's_name', 'studentname']
    for col in df.columns:
        if str(col).strip().lower() in name_variations:
            detected['name'] = col
            break
    
    # Common variations for enrollment column
    enrollment_variations = ['enrollment', 'enrollment_no', 'enrollmentno', 'enroll', 'roll_no', 'rollno', 'id']
    for col in df.columns:
        if str(col).strip().lower() in enrollment_variations:
            detected['enrollment'] = col
            break
    
    # Detect question columns (Q1, Q2, etc.)
    question_cols = []
    for col in df.columns:
        col_str = str(col).strip().lower()
        if col_str.startswith('q') and col_str[1:].isdigit():
            question_cols.append(col)
    
    detected['questions'] = sorted(question_cols, key=lambda x: int(x[1:]))
    
    return detected

def convert_to_standard_format(input_file, output_file):
    """Convert any Excel file to the standard Student PDF Generator format"""
    
    print(f"🔍 Reading input file: {input_file}")
    
    try:
        # Read the Excel file
        df = pd.read_excel(input_file, dtype=str)
        print(f"✅ Successfully read {len(df)} rows")
        
        # Detect column mapping
        detected = detect_columns(df)
        print(f"📋 Detected columns:")
        print(f"   Name column: {detected['name']}")
        print(f"   Enrollment column: {detected['enrollment']}")
        print(f"   Question columns: {len(detected['questions'])} found")
        
        if not detected['name'] or not detected['enrollment']:
            print("❌ Could not detect required columns (Name and Enrollment)")
            return False
        
        if len(detected['questions']) < 10:
            print(f"⚠️  Only {len(detected['questions'])} questions found. Minimum 10 required.")
            response = input("Continue anyway? (y/n): ")
            if response.lower() != 'y':
                return False
        
        # Create new workbook with standard format
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Create standard headers
        headers = ["Name", "EnrollmentNo"]
        num_questions = min(len(detected['questions']), 20)  # Max 20 questions
        
        for i in range(1, num_questions + 1):
            headers += [
                f"Q{i}",
                f"Q{i}_A", f"Q{i}_B", f"Q{i}_C", f"Q{i}_D",
                f"Q{i}_Answer"
            ]
        
        # Style headers
        header_fill = PatternFill("solid", fgColor="1F3A8F")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        ws.row_dimensions[1].height = 22
        
        # Process each row
        row_fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]
        valid_rows = 0
        
        for idx, (_, row) in enumerate(df.iterrows(), start=2):
            # Get name and enrollment
            name = str(row[detected['name']]).strip() if pd.notna(row[detected['name']]) else ""
            enrollment = str(row[detected['enrollment']]).strip() if pd.notna(row[detected['enrollment']]) else ""
            
            if not name or not enrollment or name.lower() == 'nan' or enrollment.lower() == 'nan':
                print(f"⚠️  Skipping row {idx}: Missing name or enrollment")
                continue
            
            # Add row to new worksheet
            row_data = [name, enrollment]
            
            # Process questions
            for i in range(num_questions):
                if i < len(detected['questions']):
                    q_col = detected['questions'][i]
                    question_text = str(row[q_col]).strip() if pd.notna(row[q_col]) else ""
                    
                    # Extract options (look for columns with this question number)
                    options = {}
                    for opt in ['A', 'B', 'C', 'D']:
                        opt_col = None
                        for col in df.columns:
                            col_str = str(col).strip().lower()
                            if col_str.startswith(f'q{i+1}_') and col_str.endswith(opt.lower()):
                                opt_col = col
                                break
                        
                        if opt_col and pd.notna(row[opt_col]):
                            options[opt] = str(row[opt_col]).strip()
                    
                    # Look for answer column
                    answer_col = None
                    for col in df.columns:
                        col_str = str(col).strip().lower()
                        if col_str.startswith(f'q{i+1}_') and 'answer' in col_str:
                            answer_col = col
                            break
                    
                    answer = str(row[answer_col]).strip().upper() if answer_col and pd.notna(row[answer_col]) else ""
                    
                    # Add to row data
                    row_data.extend([
                        question_text,  # Q{i}
                        options.get('A', ''), options.get('B', ''), 
                        options.get('C', ''), options.get('D', ''),
                        answer  # Q{i}_Answer
                    ])
                else:
                    # Add empty columns for missing questions
                    row_data.extend(['', '', '', '', '', ''])
            
            # Write row to worksheet
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=valid_rows + 2, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Alternate row colors
            ws.row_dimensions[valid_rows + 2].height = 20
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=valid_rows + 2, column=col_idx).fill = row_fills[valid_rows % 2]
            
            valid_rows += 1
        
        if valid_rows == 0:
            print("❌ No valid rows found to convert")
            return False
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            if col_idx == 1:  # Name column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            elif col_idx == 2:  # Enrollment column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Save the converted file
        wb.save(output_file)
        
        print(f"✅ Successfully converted {valid_rows} students to standard format")
        print(f"📁 Output saved to: {output_file}")
        print(f"📊 Format: {num_questions} questions per student")
        
        return True
        
    except Exception as e:
        print(f"❌ Error converting file: {e}")
        return False

def main():
    if len(sys.argv) != 3:
        print("Usage: python convert_excel.py input_file.xlsx output_file.xlsx")
        print("\nExample:")
        print("  python convert_excel.py mcq_basic_sample.xlsx converted_students.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return
    
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        print(f"❌ Input file must be Excel format (.xlsx or .xls)")
        return
    
    # Convert the file
    success = convert_to_standard_format(input_file, output_file)
    
    if success:
        print(f"\n🎉 Conversion completed successfully!")
        print(f"📋 You can now use {output_file} with the Student PDF Generator")
    else:
        print(f"\n❌ Conversion failed. Please check your input file format.")

if __name__ == "__main__":
    main()
