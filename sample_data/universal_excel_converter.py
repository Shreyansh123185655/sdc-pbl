"""
Universal Excel File Analyzer and Converter
Intelligently analyzes any Excel file structure and converts it to the required Student PDF Generator format.

Features:
- Automatic column detection and mapping
- Multiple Excel format support
- Question bank conversion
- Student data extraction
- Flexible structure recognition
- Error recovery and validation
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import sys
import os
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelAnalyzer:
    """Analyzes Excel file structure and determines the best conversion strategy"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.df = None
        self.structure_type = None
        self.column_mapping = {}
        self.analysis_result = {}
        
    def load_file(self) -> bool:
        """Load Excel file and detect structure"""
        try:
            self.df = pd.read_excel(self.file_path, dtype=str)
            logger.info(f"Successfully loaded Excel file: {len(self.df)} rows, {len(self.df.columns)} columns")
            return True
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            return False
    
    def analyze_structure(self) -> Dict[str, Any]:
        """Analyze the Excel file structure and determine conversion strategy"""
        if self.df is None:
            return {"error": "File not loaded"}
        
        analysis = {
            "total_rows": len(self.df),
            "total_columns": len(self.df.columns),
            "columns": list(self.df.columns),
            "sample_data": self.df.head(3).to_dict('records'),
            "structure_type": None,
            "confidence": 0,
            "issues": []
        }
        
        # Detect structure type
        structure_detection = self._detect_structure_type()
        analysis.update(structure_detection)
        
        self.analysis_result = analysis
        return analysis
    
    def _detect_structure_type(self) -> Dict[str, Any]:
        """Detect the type of Excel structure"""
        columns = [str(col).strip().lower() for col in self.df.columns]
        
        # Structure type detection patterns
        patterns = {
            "standard_student_format": {
                "required_columns": ["name", "enrollmentno"],
                "question_pattern": r"^q\d+$",
                "option_pattern": r"^q\d+_[a-d]$",
                "answer_pattern": r"^q\d+_answer$",
                "confidence_threshold": 0.8
            },
            "question_bank_format": {
                "required_columns": ["questiontext", "optiona", "optionb"],
                "alternative_columns": ["question", "a", "b"],
                "confidence_threshold": 0.7
            },
            "simple_student_format": {
                "required_columns": ["name", "roll"],
                "alternative_names": {"roll": "enrollmentno", "rollno": "enrollmentno"},
                "confidence_threshold": 0.6
            },
            "tabular_format": {
                "indicators": ["s.no", "serial", "question", "answer"],
                "confidence_threshold": 0.5
            }
        }
        
        best_match = {"structure_type": "unknown", "confidence": 0}
        
        for structure_type, pattern in patterns.items():
            confidence = self._calculate_pattern_match(columns, pattern)
            if confidence > best_match["confidence"]:
                best_match = {
                    "structure_type": structure_type,
                    "confidence": confidence,
                    "pattern": pattern
                }
        
        # Determine if conversion is needed
        needs_conversion = best_match["confidence"] < 0.8
        
        return {
            "structure_type": best_match["structure_type"],
            "confidence": best_match["confidence"],
            "needs_conversion": needs_conversion,
            "recommended_action": self._get_recommended_action(best_match)
        }
    
    def _calculate_pattern_match(self, columns: List[str], pattern: Dict) -> float:
        """Calculate how well columns match a pattern"""
        score = 0
        total_checks = 0
        
        # Check required columns
        if "required_columns" in pattern:
            total_checks += len(pattern["required_columns"])
            for req_col in pattern["required_columns"]:
                if req_col in columns:
                    score += 1
                # Check alternative names
                elif "alternative_names" in pattern:
                    for alt_name, target in pattern["alternative_names"].items():
                        if alt_name in columns and target == req_col:
                            score += 1
                            break
        
        # Check question patterns
        if "question_pattern" in pattern:
            question_cols = [col for col in columns if re.match(pattern["question_pattern"], col)]
            if question_cols:
                score += len(question_cols) * 0.5
                total_checks += len(question_cols) * 0.5
        
        # Check option patterns
        if "option_pattern" in pattern:
            option_cols = [col for col in columns if re.match(pattern["option_pattern"], col)]
            if option_cols:
                score += len(option_cols) * 0.3
                total_checks += len(option_cols) * 0.3
        
        # Check answer patterns
        if "answer_pattern" in pattern:
            answer_cols = [col for col in columns if re.match(pattern["answer_pattern"], col)]
            if answer_cols:
                score += len(answer_cols) * 0.2
                total_checks += len(answer_cols) * 0.2
        
        # Check indicators
        if "indicators" in pattern:
            indicator_count = sum(1 for indicator in pattern["indicators"] if indicator in columns)
            score += indicator_count * 0.1
            total_checks += len(pattern["indicators"]) * 0.1
        
        return score / total_checks if total_checks > 0 else 0
    
    def _get_recommended_action(self, best_match: Dict) -> str:
        """Get recommended action based on structure analysis"""
        confidence = best_match["confidence"]
        structure_type = best_match["structure_type"]
        
        if confidence >= 0.8:
            return "direct_use"
        elif confidence >= 0.6:
            return "auto_convert"
        elif confidence >= 0.4:
            return "manual_review"
        else:
            return "manual_conversion"


class UniversalExcelConverter:
    """Universal converter that can handle any Excel file structure"""
    
    def __init__(self):
        self.supported_formats = [
            "standard_student_format",
            "question_bank_format", 
            "simple_student_format",
            "tabular_format",
            "custom_format"
        ]
    
    def convert_to_standard_format(self, file_path: str, output_path: str, 
                                  structure_type: str = None, 
                                  custom_mapping: Dict = None) -> bool:
        """Convert any Excel file to standard Student PDF Generator format"""
        
        # Analyze the file
        analyzer = ExcelAnalyzer(file_path)
        if not analyzer.load_file():
            return False
        
        analysis = analyzer.analyze_structure()
        
        # Determine conversion strategy
        if structure_type is None:
            structure_type = analysis.get("structure_type", "unknown")
        
        logger.info(f"Detected structure: {structure_type} (confidence: {analysis.get('confidence', 0):.2f})")
        
        # Choose conversion method
        conversion_methods = {
            "standard_student_format": self._convert_standard_format,
            "question_bank_format": self._convert_question_bank,
            "simple_student_format": self._convert_simple_student,
            "tabular_format": self._convert_tabular,
            "custom_format": self._convert_custom
        }
        
        converter = conversion_methods.get(structure_type, self._convert_generic)
        
        try:
            return converter(analyzer.df, output_path, custom_mapping)
        except Exception as e:
            logger.error(f"Conversion failed: {e}")
            # Fallback to generic conversion
            return self._convert_generic(analyzer.df, output_path, custom_mapping)
    
    def _convert_standard_format(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Convert standard student format (already compatible)"""
        try:
            # Check if already in correct format
            required_cols = ['name', 'enrollmentno']
            if all(col.lower() in [str(c).lower() for c in df.columns] for col in required_cols):
                # Just save as is
                df.to_excel(output_path, index=False)
                logger.info("File already in standard format, saved directly")
                return True
            else:
                # Apply column mapping if provided
                return self._apply_column_mapping(df, output_path, custom_mapping)
        except Exception as e:
            logger.error(f"Standard format conversion failed: {e}")
            return False
    
    def _convert_question_bank(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Convert question bank format to student format"""
        try:
            # Detect question bank columns
            question_col = self._find_column(df, ['questiontext', 'question', 'qtext', 'text'])
            option_a_col = self._find_column(df, ['optiona', 'a', 'choicea', 'opta'])
            option_b_col = self._find_column(df, ['optionb', 'b', 'choiceb', 'optb'])
            option_c_col = self._find_column(df, ['optionc', 'c', 'choicec', 'optc'])
            option_d_col = self._find_column(df, ['optiond', 'd', 'choiced', 'optd'])
            answer_col = self._find_column(df, ['correctanswer', 'answer', 'correct', 'key'])
            
            if not all([question_col, option_a_col, option_b_col]):
                logger.error("Question bank format not detected properly")
                return False
            
            # Create student data from question bank
            students = []
            base_names = ["Aarav Sharma", "Priya Patel", "Rohit Verma", "Sneha Gupta", "Arjun Singh"]
            
            # Limit to reasonable number of questions
            max_questions = min(20, len(df))
            questions_df = df.head(max_questions)
            
            for i, name in enumerate(base_names[:5]):  # Create 5 students
                student_data = {
                    'Name': name,
                    'EnrollmentNo': f'EN2024{str(i+1).zfill(3)}'
                }
                
                # Add questions with potential shuffling
                questions = questions_df.copy()
                if i > 0:  # Shuffle for different students
                    questions = questions.sample(frac=1).reset_index(drop=True)
                
                for j, (_, row) in enumerate(questions.iterrows(), 1):
                    student_data[f'Q{j}'] = str(row[question_col]).strip()
                    student_data[f'Q{j}_A'] = str(row[option_a_col]).strip()
                    student_data[f'Q{j}_B'] = str(row[option_b_col]).strip()
                    student_data[f'Q{j}_C'] = str(row[option_c_col]).strip() if option_c_col else ''
                    student_data[f'Q{j}_D'] = str(row[option_d_col]).strip() if option_d_col else ''
                    student_data[f'Q{j}_Answer'] = str(row[answer_col]).strip().upper() if answer_col else ''
                
                students.append(student_data)
            
            # Create DataFrame and save
            result_df = pd.DataFrame(students)
            self._save_with_formatting(result_df, output_path)
            logger.info(f"Converted question bank to {len(students)} students with {max_questions} questions each")
            return True
            
        except Exception as e:
            logger.error(f"Question bank conversion failed: {e}")
            return False
    
    def _convert_simple_student(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Convert simple student format with basic column mapping"""
        try:
            # Map common column variations
            column_map = {
                'name': ['name', 'student', 'studentname', 's_name'],
                'enrollmentno': ['enrollmentno', 'enrollment', 'roll', 'rollno', 'roll_no', 'id']
            }
            
            mapped_df = df.copy()
            actual_mapping = {}
            
            for target_col, possible_cols in column_map.items():
                for possible_col in possible_cols:
                    for actual_col in df.columns:
                        if str(actual_col).lower() == possible_col.lower():
                            mapped_df[target_col] = df[actual_col]
                            actual_mapping[target_col] = actual_col
                            break
            
            # Check if we have the required columns
            if 'name' not in mapped_df.columns or 'enrollmentno' not in mapped_df.columns:
                logger.error("Could not map required columns for simple student format")
                return False
            
            # Process questions (assume they exist in some format)
            question_cols = [col for col in df.columns if re.match(r'^q\d+', str(col).lower(), re.IGNORECASE)]
            
            if not question_cols:
                logger.warning("No question columns found, creating placeholder questions")
                # Add placeholder questions
                for i in range(1, 11):  # Add 10 placeholder questions
                    mapped_df[f'Q{i}'] = f"Question {i} (Please edit)"
                    mapped_df[f'Q{i}_A'] = "Option A"
                    mapped_df[f'Q{i}_B'] = "Option B"
                    mapped_df[f'Q{i}_C'] = "Option C"
                    mapped_df[f'Q{i}_D'] = "Option D"
                    mapped_df[f'Q{i}_Answer'] = "A"
            
            self._save_with_formatting(mapped_df, output_path)
            logger.info(f"Converted simple student format with {len(question_cols)} questions")
            return True
            
        except Exception as e:
            logger.error(f"Simple student conversion failed: {e}")
            return False
    
    def _convert_tabular(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Convert tabular format to student format"""
        try:
            # This is more complex - try to infer structure
            # Look for patterns that suggest student vs question data
            
            # If first column looks like student names
            first_col = df.columns[0]
            if self._looks_like_student_names(df[first_col]):
                return self._convert_from_tabular_student_data(df, output_path)
            else:
                # Try to treat as question bank
                return self._convert_question_bank(df, output_path)
                
        except Exception as e:
            logger.error(f"Tabular conversion failed: {e}")
            return False
    
    def _convert_custom(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Convert using custom mapping provided by user"""
        if not custom_mapping:
            logger.error("Custom mapping required for custom format conversion")
            return False
        
        try:
            mapped_df = pd.DataFrame()
            
            for target_col, source_col in custom_mapping.items():
                if source_col in df.columns:
                    mapped_df[target_col] = df[source_col]
                else:
                    logger.warning(f"Source column '{source_col}' not found")
            
            self._save_with_formatting(mapped_df, output_path)
            logger.info("Custom format conversion completed")
            return True
            
        except Exception as e:
            logger.error(f"Custom conversion failed: {e}")
            return False
    
    def _convert_generic(self, df: pd.DataFrame, output_path: str, custom_mapping: Dict = None) -> bool:
        """Generic conversion as last resort"""
        try:
            logger.info("Attempting generic conversion...")
            
            # Create a basic structure with available data
            result_df = pd.DataFrame()
            
            # Try to find name column
            name_col = self._find_column(df, ['name', 'student', 'person', 'user'])
            if name_col:
                result_df['Name'] = df[name_col]
            else:
                # Create generic names
                result_df['Name'] = [f"Student {i+1}" for i in range(len(df))]
            
            # Try to find ID column
            id_col = self._find_column(df, ['id', 'roll', 'enrollment', 'number'])
            if id_col:
                result_df['EnrollmentNo'] = df[id_col]
            else:
                # Generate IDs
                result_df['EnrollmentNo'] = [f"STU{str(i+1).zfill(4)}" for i in range(len(df))]
            
            # Add placeholder questions
            for i in range(1, 11):  # 10 questions
                result_df[f'Q{i}'] = f"Question {i} (Please edit with actual content)"
                result_df[f'Q{i}_A'] = "Option A"
                result_df[f'Q{i}_B'] = "Option B" 
                result_df[f'Q{i}_C'] = "Option C"
                result_df[f'Q{i}_D'] = "Option D"
                result_df[f'Q{i}_Answer'] = "A"
            
            self._save_with_formatting(result_df, output_path)
            logger.info("Generic conversion completed - please review and edit the generated file")
            return True
            
        except Exception as e:
            logger.error(f"Generic conversion failed: {e}")
            return False
    
    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """Find the best matching column from possible names"""
        for possible_name in possible_names:
            for col in df.columns:
                if str(col).strip().lower() == possible_name.lower():
                    return col
        return None
    
    def _looks_like_student_names(self, series: pd.Series) -> bool:
        """Check if a series looks like it contains student names"""
        # Simple heuristic: check first few values
        sample = series.head(10).dropna()
        if len(sample) == 0:
            return False
        
        # Check if values look like names (contain letters, reasonable length)
        name_like_count = 0
        for value in sample:
            val_str = str(value).strip()
            if (len(val_str) > 3 and len(val_str) < 50 and 
                any(c.isalpha() for c in val_str) and
                not any(c.isdigit() for c in val_str.split()[0] if val_str.split())):
                name_like_count += 1
        
        return name_like_count / len(sample) > 0.6
    
    def _convert_from_tabular_student_data(self, df: pd.DataFrame, output_path: str) -> bool:
        """Convert tabular data where first column is student names"""
        try:
            # Assume first column is student names
            student_names = df.iloc[:, 0].dropna()
            
            # Try to extract questions from other columns
            question_data = []
            for i, name in enumerate(student_names):
                student_data = {'Name': name, 'EnrollmentNo': f'STU{str(i+1).zfill(4)}'}
                
                # Process remaining columns as questions
                col_idx = 1
                question_num = 1
                
                while col_idx < len(df.columns) and question_num <= 20:
                    # Try to extract question and options from consecutive columns
                    if col_idx + 3 < len(df.columns):
                        student_data[f'Q{question_num}'] = str(df.iloc[i, col_idx]).strip()
                        student_data[f'Q{question_num}_A'] = str(df.iloc[i, col_idx + 1]).strip()
                        student_data[f'Q{question_num}_B'] = str(df.iloc[i, col_idx + 2]).strip()
                        student_data[f'Q{question_num}_C'] = str(df.iloc[i, col_idx + 3]).strip()
                        student_data[f'Q{question_num}_D'] = ""
                        student_data[f'Q{question_num}_Answer'] = "A"
                        col_idx += 4
                    else:
                        # Not enough columns for full question, add placeholder
                        student_data[f'Q{question_num}'] = f"Question {question_num}"
                        student_data[f'Q{question_num}_A'] = "Option A"
                        student_data[f'Q{question_num}_B'] = "Option B"
                        student_data[f'Q{question_num}_C'] = "Option C"
                        student_data[f'Q{question_num}_D'] = "Option D"
                        student_data[f'Q{question_num}_Answer'] = "A"
                        col_idx += 1
                    
                    question_num += 1
                
                question_data.append(student_data)
            
            result_df = pd.DataFrame(question_data)
            self._save_with_formatting(result_df, output_path)
            logger.info(f"Converted tabular student data for {len(question_data)} students")
            return True
            
        except Exception as e:
            logger.error(f"Tabular student data conversion failed: {e}")
            return False
    
    def _apply_column_mapping(self, df: pd.DataFrame, output_path: str, mapping: Dict) -> bool:
        """Apply custom column mapping"""
        try:
            if not mapping:
                return False
            
            mapped_df = pd.DataFrame()
            for target_col, source_col in mapping.items():
                if source_col in df.columns:
                    mapped_df[target_col] = df[source_col]
            
            self._save_with_formatting(mapped_df, output_path)
            return True
            
        except Exception as e:
            logger.error(f"Column mapping failed: {e}")
            return False
    
    def _save_with_formatting(self, df: pd.DataFrame, output_path: str) -> None:
        """Save DataFrame with proper Excel formatting"""
        try:
            # Create workbook with formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Write headers
            headers = list(df.columns)
            header_fill = PatternFill("solid", fgColor="1F3A8F")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            ws.row_dimensions[1].height = 22
            
            # Write data
            row_fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]
            
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Alternate row colors
                ws.row_dimensions[row_idx].height = 20
                fill = row_fills[(row_idx - 2) % 2]
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
            
            # Adjust column widths
            for col_idx in range(1, len(headers) + 1):
                if col_idx == 1:  # Name column
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
                elif col_idx == 2:  # Enrollment column
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
                else:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"File saved with formatting: {output_path}")
            
        except Exception as e:
            logger.error(f"Failed to save with formatting: {e}")
            # Fallback to simple save
            df.to_excel(output_path, index=False)


def analyze_and_convert_excel(input_file: str, output_file: str = None, 
                              force_convert: bool = False) -> Dict[str, Any]:
    """
    Analyze and convert any Excel file to the required format
    
    Args:
        input_file: Path to input Excel file (string path or UploadFile object)
        output_file: Path to output file (auto-generated if None)
        force_convert: Force conversion even if file is already compatible
    
    Returns:
        Dictionary with analysis results and conversion status
    """
    
    # Handle both file path string and UploadFile object
    if hasattr(input_file, 'file'):
        # It's an UploadFile object - read bytes
        file_bytes = input_file.file.read()
        input_filename = input_file.filename
        logger.info(f"Processing UploadFile object: {input_filename}")
    else:
        # It's a file path string
        if not os.path.exists(input_file):
            return {"error": "Input file not found"}
        
        file_bytes = open(input_file, 'rb').read()
        input_filename = os.path.basename(input_file)
        logger.info(f"Processing file: {input_filename}")
    
    # Generate output filename if not provided
    if output_file is None:
        input_path = Path(input_filename if hasattr(input_file, 'file') else input_file)
        output_file = input_path.parent / f"converted_{input_path.stem}.xlsx"
    
    # Initialize converter
    converter = UniversalExcelConverter()
    
    # Analyze file
    analyzer = ExcelAnalyzer(input_filename if hasattr(input_file, 'file') else input_file)
    if not analyzer.load_file():
        return {"error": "Failed to load Excel file"}
    
    analysis = analyzer.analyze_structure()
    
    # Determine if conversion is needed
    needs_conversion = analysis.get("needs_conversion", True) or force_convert
    
    result = {
        "analysis": analysis,
        "needs_conversion": needs_conversion,
        "input_file": input_filename,
        "output_file": str(output_file)
    }
    
    if needs_conversion:
        # Perform conversion
        success = converter.convert_to_standard_format(
            input_filename if hasattr(input_file, 'file') else input_file, 
            str(output_file), 
            custom_mapping=None
        )
        result["conversion_success"] = success
        
        if success:
            result["message"] = f"File converted successfully and saved to {output_file}"
        else:
            result["message"] = "Conversion failed - please check logs for details"
    else:
        # File is already compatible
        result["conversion_success"] = True
        result["message"] = "File is already in compatible format - no conversion needed"
    
    return result


def main():
    """Command line interface"""
    if len(sys.argv) < 2:
        print("Usage: python universal_excel_converter.py <input_file> [output_file] [--force]")
        print("Example: python universal_excel_converter.py my_data.xlsx converted.xlsx")
        print("Example: python universal_excel_converter.py my_data.xlsx --force")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith("--") else None
    force_convert = "--force" in sys.argv
    
    result = analyze_and_convert_excel(input_file, output_file, force_convert)
    
    if "error" in result:
        print(f"❌ Error: {result['error']}")
        return
    
    # Display results
    print("🔍 Excel File Analysis Results:")
    print("=" * 50)
    print(f"📁 Input file: {result['input_file']}")
    print(f"📊 Structure type: {result['analysis']['structure_type']}")
    print(f"🎯 Confidence: {result['analysis']['confidence']:.2f}")
    print(f"📋 Total rows: {result['analysis']['total_rows']}")
    print(f"📋 Total columns: {result['analysis']['total_columns']}")
    print(f"🔄 Needs conversion: {'Yes' if result['needs_conversion'] else 'No'}")
    print()
    
    if result['needs_conversion']:
        print(f"📁 Output file: {result['output_file']}")
        print(f"✅ Conversion status: {'Success' if result['conversion_success'] else 'Failed'}")
        print(f"💬 Message: {result['message']}")
    else:
        print(f"💬 Message: {result['message']}")
    
    # Show column information
    print("\n📋 Detected Columns:")
    for i, col in enumerate(result['analysis']['columns'], 1):
        print(f"  {i:2d}. {col}")
    
    # Show sample data
    print("\n📋 Sample Data (first 2 rows):")
    sample_data = result['analysis']['sample_data'][:2]
    for i, row in enumerate(sample_data, 1):
        print(f"  Row {i}:")
        for key, value in list(row.items())[:5]:  # Show first 5 columns
            print(f"    {key}: {value}")
        if len(row) > 5:
            print(f"    ... and {len(row) - 5} more columns")
        print()


if __name__ == "__main__":
    main()
