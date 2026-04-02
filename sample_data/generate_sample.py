"""
Run this script once to generate a sample Excel file for testing.
Usage: python generate_sample.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Students"

# ── Header row ──────────────────────────────────────────────────────────────
headers = ["Name", "EnrollmentNo"]
for i in range(1, 16):          # 15 questions
    headers += [
        f"Q{i}",
        f"Q{i}_A", f"Q{i}_B", f"Q{i}_C", f"Q{i}_D",
        f"Q{i}_Answer"
    ]

header_fill = PatternFill("solid", fgColor="1F3A8F")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

ws.row_dimensions[1].height = 22

# ── Student data ─────────────────────────────────────────────────────────────
students = [
    ("Aarav Sharma",   "EN2024001"),
    ("Priya Patel",    "EN2024002"),
    ("Rohit Verma",    "EN2024003"),
    ("Sneha Gupta",    "EN2024004"),
    ("Arjun Singh",    "EN2024005"),
]

QUESTIONS = [
    ("What is the full form of CPU?",
     "Central Processing Unit", "Central Program Unit",
     "Computer Processing Unit", "Core Processing Unit", "A"),

    ("Which data structure uses LIFO?",
     "Queue", "Stack", "Tree", "Graph", "B"),

    ("What does HTTP stand for?",
     "HyperText Transfer Protocol", "High Transfer Text Protocol",
     "Hyper Text Transmission Protocol", "HyperText Transmission Port", "A"),

    ("Which language is used in Android development?",
     "Swift", "Kotlin", "Ruby", "Go", "B"),

    ("What is the time complexity of binary search?",
     "O(n)", "O(n²)", "O(log n)", "O(1)", "C"),

    ("Which of the following is NOT an OOP concept?",
     "Encapsulation", "Polymorphism", "Compilation", "Inheritance", "C"),

    ("What does SQL stand for?",
     "Structured Query Language", "Simple Query Language",
     "Structured Question Language", "Sequential Query List", "A"),

    ("Which HTML tag is used for the largest heading?",
     "<h6>", "<h1>", "<head>", "<header>", "B"),

    ("What is Git primarily used for?",
     "Database management", "Version control",
     "Web hosting", "Containerization", "B"),

    ("Which protocol is used to send emails?",
     "FTP", "HTTP", "SMTP", "SSH", "C"),

    ("What does RAM stand for?",
     "Read Access Memory", "Random Access Memory",
     "Rapid Access Memory", "Read Assign Memory", "B"),

    ("Which of these is a NoSQL database?",
     "MySQL", "PostgreSQL", "MongoDB", "SQLite", "C"),

    ("What is the default port for HTTPS?",
     "80", "8080", "443", "22", "C"),

    ("Which CSS property controls text size?",
     "font-weight", "text-size", "font-size", "letter-spacing", "C"),

    ("What does API stand for?",
     "Application Programming Interface",
     "Application Program Integration",
     "Applied Programming Interface",
     "Application Protocol Interface", "A"),
]

row_fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]

for s_idx, (name, enroll) in enumerate(students):
    row = s_idx + 2
    data = [name, enroll]
    for q_text, a, b, c, d, ans in QUESTIONS:
        data += [q_text, a, b, c, d, ans]

    fill = row_fills[s_idx % 2]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    ws.row_dimensions[row].height = 18

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 16
for col_idx in range(3, len(headers) + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 28

wb.save("students_sample.xlsx")
print("✅  students_sample.xlsx created successfully!")
