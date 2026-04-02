"""
Question Bank to Student Format Converter
Converts question bank format to student-based format for Student PDF Generator.

Usage: python convert_question_bank.py question_bank.xlsx output.xlsx [num_students]
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import random

def convert_question_bank_to_students(input_file, output_file, num_students=5):
    """Convert question bank to student format"""
    
    print(f"🔍 Reading question bank: {input_file}")
    
    try:
        # Read the question bank
        df = pd.read_excel(input_file, dtype=str)
        print(f"✅ Successfully read {len(df)} questions")
        
        # Validate required columns
        required_cols = ['QuestionText', 'OptionA', 'OptionB', 'OptionC', 'OptionD', 'CorrectAnswer']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            print(f"❌ Missing required columns: {missing_cols}")
            return False
        
        # Filter out questions with missing data
        valid_questions = []
        for idx, row in df.iterrows():
            question_text = str(row['QuestionText']).strip()
            if question_text and question_text.lower() != 'nan':
                valid_questions.append(row)
        
        print(f"✅ Found {len(valid_questions)} valid questions")
        
        if len(valid_questions) < 10:
            print(f"❌ Need at least 10 questions, found only {len(valid_questions)}")
            return False
        
        # Create student data
        students = [
            ("Aarav Sharma", "EN2024001"),
            ("Priya Patel", "EN2024002"),
            ("Rohit Verma", "EN2024003"),
            ("Sneha Gupta", "EN2024004"),
            ("Arjun Singh", "EN2024005"),
            ("Kavya Reddy", "EN2024006"),
            ("Rahul Kumar", "EN2024007"),
            ("Anjali Singh", "EN2024008"),
            ("Vikram Malhotra", "EN2024009"),
            ("Divya Sharma", "EN2024010"),
        ]
        
        # Limit to requested number of students
        students = students[:min(num_students, len(students))]
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Determine number of questions (max 20)
        num_q_to_use = min(len(valid_questions), 20)
        
        # Create headers
        headers = ["Name", "EnrollmentNo"]
        for i in range(1, num_q_to_use + 1):
            headers += [
                f"Q{i}", f"Q{i}_A", f"Q{i}_B", f"Q{i}_C", f"Q{i}_D", f"Q{i}_Answer"
            ]
        
        # Style headers
        header_fill = PatternFill("solid", fgColor="1F3A8F")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        ws.row_dimensions[1].height = 22
        
        # Add student data
        row_fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]
        
        for student_idx, (name, enrollment) in enumerate(students):
            row_data = [name, enrollment]
            
            # Add questions (shuffle order for each student if desired)
            questions_to_use = valid_questions[:num_q_to_use]
            
            for q_idx, question in enumerate(questions_to_use):
                row_data.extend([
                    str(question['QuestionText']).strip(),
                    str(question['OptionA']).strip(),
                    str(question['OptionB']).strip(),
                    str(question['OptionC']).strip(),
                    str(question['OptionD']).strip(),
                    str(question['CorrectAnswer']).strip().upper()
                ])
            
            # Write row to worksheet
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=student_idx + 2, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Alternate row colors
            ws.row_dimensions[student_idx + 2].height = 20
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=student_idx + 2, column=col_idx).fill = row_fills[student_idx % 2]
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            if col_idx == 1:  # Name column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            elif col_idx == 2:  # Enrollment column
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Save the file
        wb.save(output_file)
        
        print(f"✅ Successfully created {len(students)} student records")
        print(f"📊 Each student has {num_q_to_use} questions")
        print(f"📁 Output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error converting file: {e}")
        return False

def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_question_bank.py question_bank.xlsx output.xlsx [num_students]")
        print("\nExample:")
        print("  python convert_question_bank.py mcq_basic_sample.xlsx students.xlsx 5")
        print("\nDefault: Creates 5 students if not specified")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    num_students = int(sys.argv[3]) if len(sys.argv) > 3 else 5
    
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return
    
    success = convert_question_bank_to_students(input_file, output_file, num_students)
    
    if success:
        print(f"\n🎉 Conversion completed successfully!")
        print(f"📋 You can now use {output_file} with the Student PDF Generator")
        print(f"👥 Created {num_students} students with the same set of questions")
    else:
        print(f"\n❌ Conversion failed.")

if __name__ == "__main__":
    import os
    main()
